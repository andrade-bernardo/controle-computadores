import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from pathlib import Path
import sys
import shutil

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

DB_PATH = BASE_DIR / "computadores.db"
BACKUP_DIR = BASE_DIR / "backups"
RELATORIOS_DIR = BASE_DIR / "relatorios"

STATUS = ["Em uso", "Estoque", "Manutenção", "Baixado"]

THEME = {
    "light": {
        "bg_main":        "#F0F2F5",
        "bg_card":        "#FFFFFF",
        "bg_sidebar":     "#0F172A",
        "bg_sidebar_sel": "#1E293B",
        "fg_title":       "#0F172A",
        "fg_body":        "#374151",
        "fg_muted":       "#6B7280",
        "fg_sidebar":     "#94A3B8",
        "fg_sidebar_sel": "#F8FAFC",
        "accent":         "#2563EB",
        "accent_hover":   "#1D4ED8",
        "danger":         "#DC2626",
        "danger_hover":   "#B91C1C",
        "success":        "#059669",
        "warning":        "#D97706",
        "purple":         "#7C3AED",
        "border":         "#E2E8F0",
        "row_even":       "#FFFFFF",
        "row_odd":        "#F8FAFC",
        "row_sel":        "#2563EB",
        "tree_bg":        "#FFFFFF",
        "tree_head_bg":   "#F1F5F9",
        "tree_head_fg":   "#0F172A",
        "btn_neutral_bg": "#E2E8F0",
        "btn_neutral_fg": "#0F172A",
        "entry_border":   "#CBD5E1",
        "toast_bg":       "#059669",
        "toast_fg":       "#FFFFFF",
    },
    "dark": {
        "bg_main":        "#0D1117",
        "bg_card":        "#161B22",
        "bg_sidebar":     "#010409",
        "bg_sidebar_sel": "#1C2128",
        "fg_title":       "#F0F6FC",
        "fg_body":        "#CDD9E5",
        "fg_muted":       "#768390",
        "fg_sidebar":     "#636E7B",
        "fg_sidebar_sel": "#F0F6FC",
        "accent":         "#388BFD",
        "accent_hover":   "#1F6FEB",
        "danger":         "#F85149",
        "danger_hover":   "#DA3633",
        "success":        "#3FB950",
        "warning":        "#D29922",
        "purple":         "#BC8CFF",
        "border":         "#30363D",
        "row_even":       "#161B22",
        "row_odd":        "#1C2128",
        "row_sel":        "#1F6FEB",
        "tree_bg":        "#161B22",
        "tree_head_bg":   "#1C2128",
        "tree_head_fg":   "#F0F6FC",
        "btn_neutral_bg": "#21262D",
        "btn_neutral_fg": "#CDD9E5",
        "entry_border":   "#30363D",
        "toast_bg":       "#3FB950",
        "toast_fg":       "#0D1117",
    }
}

ROWS_PER_PAGE = 15


def conectar():
    return sqlite3.connect(DB_PATH)


def criar_banco():
    with conectar() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS computadores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patrimonio TEXT UNIQUE NOT NULL,
                marca TEXT,
                modelo TEXT,
                serial TEXT,
                responsavel TEXT,
                setor TEXT,
                localizacao TEXT,
                status TEXT,
                observacoes TEXT,
                atualizado_em TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS historico (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                computador_id INTEGER,
                data_hora TEXT,
                descricao TEXT
            )
        """)


def fazer_backup_automatico():
    if not DB_PATH.exists():
        return
    BACKUP_DIR.mkdir(exist_ok=True)
    hoje = datetime.now().strftime("%Y-%m-%d")
    backup_path = BACKUP_DIR / f"backup_computadores_{hoje}.db"
    if backup_path.exists():
        return
    try:
        shutil.copy2(DB_PATH, backup_path)
    except Exception as erro:
        messagebox.showwarning("Aviso de backup", f"Não foi possível criar o backup automático.\n\nErro: {erro}")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.mode = "light"
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title("Controle de Ativos de TI")
        self.geometry("1400x820")
        self.minsize(1200, 720)

        criar_banco()
        fazer_backup_automatico()

        self.id_atual = None
        self.campos = {}
        self.filtro_status = tk.StringVar(value="Todos")
        self.pagina_atual = 0
        self.dados_cache = []

        self._toast_job = None
        self._widget_refs = {}

        self.configurar_estilos()
        self.montar_tela()
        self.carregar()
        self.atualizar_resumo()

    # ─── helpers ──────────────────────────────────────────────────────────────

    def T(self):
        return THEME[self.mode]

    def configurar_estilos(self):
        t = self.T()
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background=t["tree_bg"],
            foreground=t["fg_body"],
            rowheight=36,
            fieldbackground=t["tree_bg"],
            borderwidth=0,
            font=("Courier New", 10),
        )
        style.configure(
            "Treeview.Heading",
            background=t["tree_head_bg"],
            foreground=t["tree_head_fg"],
            font=("Courier New", 10, "bold"),
            relief="flat",
        )
        style.map(
            "Treeview",
            background=[("selected", t["row_sel"])],
            foreground=[("selected", "#FFFFFF")],
        )

    # ─── layout ───────────────────────────────────────────────────────────────

    def montar_tela(self):
        t = self.T()
        self.configure(fg_color=t["bg_main"])
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=t["bg_sidebar"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self._widget_refs["sidebar"] = self.sidebar

        self.main = ctk.CTkFrame(self, fg_color=t["bg_main"], corner_radius=0)
        self.main.grid(row=0, column=1, sticky="nsew")
        self.main.grid_columnconfigure(0, weight=1)
        self.main.grid_rowconfigure(4, weight=1)
        self._widget_refs["main"] = self.main

        self.montar_menu_lateral()
        self.montar_cabecalho()
        self.montar_cards()
        self.montar_area_busca()
        self.montar_filtros()
        self.montar_conteudo()
        self.montar_toast()

    def montar_menu_lateral(self):
        t = self.T()
        sb = self.sidebar

        logo = ctk.CTkLabel(
            sb,
            text="[ ATIVOS.TI ]",
            font=("Courier New", 17, "bold"),
            text_color="#388BFD",
        )
        logo.pack(anchor="w", padx=20, pady=(28, 2))

        sub = ctk.CTkLabel(
            sb,
            text="controle patrimonial",
            font=("Courier New", 11),
            text_color=t["fg_sidebar"],
        )
        sub.pack(anchor="w", padx=20, pady=(0, 24))

        itens = [
            ("Computadores", "▸", True),
            ("Relatórios", "▸", False),
            ("Backups", "▸", False),
            ("Configurações", "▸", False),
        ]

        for texto, icone, ativo in itens:
            cor_bg = t["bg_sidebar_sel"] if ativo else "transparent"
            cor_fg = t["fg_sidebar_sel"] if ativo else t["fg_sidebar"]
            border = t["accent"] if ativo else "transparent"

            item = ctk.CTkFrame(sb, fg_color=cor_bg, corner_radius=6)
            item.pack(fill="x", padx=12, pady=2)

            lbl = ctk.CTkLabel(
                item,
                text=f"  {icone}  {texto}",
                font=("Courier New", 13, "bold" if ativo else "normal"),
                text_color=cor_fg,
                anchor="w",
            )
            lbl.pack(fill="x", padx=8, pady=9)

        # Toggle dark/light
        sep = ctk.CTkFrame(sb, fg_color=t["border"], height=1)
        sep.pack(fill="x", padx=16, pady=(20, 14))

        toggle_frame = ctk.CTkFrame(sb, fg_color="transparent")
        toggle_frame.pack(fill="x", padx=16)

        self.toggle_label = ctk.CTkLabel(
            toggle_frame,
            text="● LIGHT MODE" if self.mode == "light" else "○ DARK MODE",
            font=("Courier New", 11),
            text_color=t["fg_sidebar"],
            anchor="w",
        )
        self.toggle_label.pack(side="left")

        self.btn_toggle = ctk.CTkButton(
            toggle_frame,
            text="◐",
            width=32,
            height=28,
            corner_radius=6,
            fg_color=t["bg_sidebar_sel"],
            hover_color=t["accent"],
            text_color=t["fg_sidebar_sel"],
            font=("Courier New", 14, "bold"),
            command=self.alternar_tema,
        )
        self.btn_toggle.pack(side="right")

        rodape = ctk.CTkLabel(
            sb,
            text="SQLite · Excel · Backup\nsistema local",
            font=("Courier New", 10),
            text_color=t["fg_sidebar"],
            justify="left",
        )
        rodape.pack(side="bottom", anchor="w", padx=20, pady=20)

    def montar_cabecalho(self):
        t = self.T()
        header = ctk.CTkFrame(self.main, fg_color=t["bg_main"], corner_radius=0)
        header.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 6))
        header.grid_columnconfigure(0, weight=1)
        self._widget_refs["header"] = header

        titulo = ctk.CTkLabel(
            header,
            text="Controle de Computadores",
            font=("Courier New", 22, "bold"),
            text_color=t["fg_title"],
        )
        titulo.grid(row=0, column=0, sticky="w")

        subtitulo = ctk.CTkLabel(
            header,
            text="Gerencie patrimônio · responsáveis · setores · localização · histórico",
            font=("Courier New", 12),
            text_color=t["fg_muted"],
        )
        subtitulo.grid(row=1, column=0, sticky="w", pady=(2, 0))

        self.label_data = ctk.CTkLabel(
            header,
            text=datetime.now().strftime("%d/%m/%Y"),
            font=("Courier New", 12, "bold"),
            text_color=t["fg_muted"],
        )
        self.label_data.grid(row=0, column=1, sticky="e", rowspan=2)

    def montar_cards(self):
        t = self.T()
        cf = ctk.CTkFrame(self.main, fg_color=t["bg_main"], corner_radius=0)
        cf.grid(row=1, column=0, sticky="ew", padx=24, pady=(6, 8))
        self._widget_refs["cards_frame"] = cf

        for i in range(4):
            cf.grid_columnconfigure(i, weight=1)

        self.card_total     = self._criar_card(cf, "TOTAL",      "0", t["accent"],  0)
        self.card_uso       = self._criar_card(cf, "EM USO",     "0", t["success"], 1)
        self.card_estoque   = self._criar_card(cf, "ESTOQUE",    "0", t["purple"],  2)
        self.card_manutencao= self._criar_card(cf, "MANUTENÇÃO", "0", t["warning"], 3)

    def _criar_card(self, parent, titulo, valor, cor, col):
        t = self.T()
        card = ctk.CTkFrame(parent, fg_color=t["bg_card"], corner_radius=12,
                            border_width=1, border_color=t["border"])
        card.grid(row=0, column=col, sticky="ew", padx=5, ipady=4)

        top = ctk.CTkFrame(card, fg_color="transparent")
        top.pack(fill="x", padx=14, pady=(12, 2))

        dot = ctk.CTkLabel(top, text="■", font=("Courier New", 12),
                           text_color=cor, anchor="w")
        dot.pack(side="left")

        lbl_titulo = ctk.CTkLabel(top, text=titulo, font=("Courier New", 11, "bold"),
                                  text_color=t["fg_muted"], anchor="w")
        lbl_titulo.pack(side="left", padx=(5, 0))

        lbl_val = ctk.CTkLabel(card, text=valor, font=("Courier New", 28, "bold"),
                               text_color=t["fg_title"], anchor="w")
        lbl_val.pack(anchor="w", padx=16, pady=(0, 12))

        return lbl_val

    def montar_area_busca(self):
        t = self.T()
        bf = ctk.CTkFrame(self.main, fg_color=t["bg_card"], corner_radius=12,
                          border_width=1, border_color=t["border"])
        bf.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 6))
        bf.grid_columnconfigure(0, weight=1)
        self._widget_refs["busca_frame"] = bf

        self.busca = tk.StringVar()

        self.entry_busca = ctk.CTkEntry(
            bf,
            textvariable=self.busca,
            placeholder_text="Pesquisar por patrimônio, marca, modelo, responsável, setor, local ou status...",
            height=40,
            corner_radius=8,
            border_color=t["entry_border"],
            font=("Courier New", 12),
            fg_color=t["bg_main"],
            text_color=t["fg_body"],
        )
        self.entry_busca.grid(row=0, column=0, sticky="ew", padx=(14, 8), pady=12)
        self.entry_busca.bind("<KeyRelease>", lambda e: self._reset_pagina())

        self.btn_limpar = ctk.CTkButton(
            bf, text="Limpar", height=40, width=90, corner_radius=8,
            fg_color=t["btn_neutral_bg"], hover_color=t["border"],
            text_color=t["btn_neutral_fg"], font=("Courier New", 12),
            command=self.limpar_busca,
        )
        self.btn_limpar.grid(row=0, column=1, padx=6, pady=12)

        self.btn_excel = ctk.CTkButton(
            bf, text="↓ Excel", height=40, width=110, corner_radius=8,
            fg_color=t["accent"], hover_color=t["accent_hover"],
            text_color="#FFFFFF", font=("Courier New", 12, "bold"),
            command=self.exportar_excel,
        )
        self.btn_excel.grid(row=0, column=2, padx=(6, 14), pady=12)

    def montar_filtros(self):
        t = self.T()
        ff = ctk.CTkFrame(self.main, fg_color=t["bg_main"], corner_radius=0)
        ff.grid(row=3, column=0, sticky="ew", padx=24, pady=(0, 6))
        self._widget_refs["filtros_frame"] = ff

        ctk.CTkLabel(ff, text="Filtrar:", font=("Courier New", 12, "bold"),
                     text_color=t["fg_muted"]).pack(side="left", padx=(0, 10))

        opcoes = ["Todos"] + STATUS
        self.btns_filtro = {}

        for op in opcoes:
            cor_map = {
                "Todos":      t["accent"],
                "Em uso":     t["success"],
                "Estoque":    t["purple"],
                "Manutenção": t["warning"],
                "Baixado":    t["danger"],
            }
            cor = cor_map.get(op, t["accent"])
            btn = ctk.CTkButton(
                ff,
                text=op,
                height=30,
                corner_radius=6,
                fg_color=cor if op == "Todos" else t["btn_neutral_bg"],
                hover_color=cor,
                text_color="#FFFFFF" if op == "Todos" else t["btn_neutral_fg"],
                font=("Courier New", 11, "bold"),
                command=lambda o=op, c=cor: self._aplicar_filtro(o, c),
            )
            btn.pack(side="left", padx=3)
            self.btns_filtro[op] = (btn, cor)

        # label contagem
        self.label_contagem = ctk.CTkLabel(
            ff, text="", font=("Courier New", 11), text_color=t["fg_muted"]
        )
        self.label_contagem.pack(side="right", padx=6)

    def montar_conteudo(self):
        t = self.T()
        conteudo = ctk.CTkFrame(self.main, fg_color=t["bg_main"], corner_radius=0)
        conteudo.grid(row=4, column=0, sticky="nsew", padx=24, pady=(0, 20))
        conteudo.grid_columnconfigure(0, weight=2)
        conteudo.grid_columnconfigure(1, weight=1)
        conteudo.grid_rowconfigure(0, weight=1)
        self._widget_refs["conteudo"] = conteudo

        tabela_card = ctk.CTkFrame(conteudo, fg_color=t["bg_card"], corner_radius=12,
                                   border_width=1, border_color=t["border"])
        tabela_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tabela_card.grid_columnconfigure(0, weight=1)
        tabela_card.grid_rowconfigure(1, weight=1)
        self._widget_refs["tabela_card"] = tabela_card

        titulo_tab = ctk.CTkLabel(tabela_card, text="Equipamentos cadastrados",
                                  font=("Courier New", 15, "bold"), text_color=t["fg_title"],
                                  anchor="w")
        titulo_tab.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))

        tabela_container = ctk.CTkFrame(tabela_card, fg_color=t["bg_card"])
        tabela_container.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 4))
        tabela_container.grid_columnconfigure(0, weight=1)
        tabela_container.grid_rowconfigure(0, weight=1)

        colunas = ("patrimonio", "marca", "modelo", "responsavel", "setor", "localizacao", "status")
        self.tabela = ttk.Treeview(tabela_container, columns=colunas, show="headings", selectmode="browse")

        titulos  = {"patrimonio":"Patrimônio","marca":"Marca","modelo":"Modelo",
                    "responsavel":"Responsável","setor":"Setor","localizacao":"Localização","status":"Status"}
        larguras = {"patrimonio":115,"marca":105,"modelo":130,"responsavel":165,
                    "setor":115,"localizacao":145,"status":105}

        for col in colunas:
            self.tabela.heading(col, text=titulos[col],
                                command=lambda c=col: self._ordenar(c))
            self.tabela.column(col, width=larguras[col], minwidth=70)

        sy = ttk.Scrollbar(tabela_container, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=sy.set)
        self.tabela.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        self.tabela.bind("<<TreeviewSelect>>", self.selecionar)

        # Paginação
        pag_frame = ctk.CTkFrame(tabela_card, fg_color=t["bg_card"])
        pag_frame.grid(row=2, column=0, sticky="ew", padx=14, pady=(4, 12))
        self._widget_refs["pag_frame"] = pag_frame

        self.btn_prev = ctk.CTkButton(pag_frame, text="← Anterior", width=100, height=28,
                                      corner_radius=6, fg_color=t["btn_neutral_bg"],
                                      hover_color=t["border"], text_color=t["btn_neutral_fg"],
                                      font=("Courier New", 11), command=self._pag_prev)
        self.btn_prev.pack(side="left", padx=(0, 8))

        self.label_pag = ctk.CTkLabel(pag_frame, text="Página 1 / 1",
                                      font=("Courier New", 11), text_color=t["fg_muted"])
        self.label_pag.pack(side="left")

        self.btn_next = ctk.CTkButton(pag_frame, text="Próxima →", width=100, height=28,
                                      corner_radius=6, fg_color=t["btn_neutral_bg"],
                                      hover_color=t["border"], text_color=t["btn_neutral_fg"],
                                      font=("Courier New", 11), command=self._pag_next)
        self.btn_next.pack(side="left", padx=8)

        self.montar_formulario(conteudo)

    def montar_formulario(self, parent):
        t = self.T()
        form = ctk.CTkFrame(parent, fg_color=t["bg_card"], corner_radius=12,
                            border_width=1, border_color=t["border"])
        form.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        form.grid_columnconfigure(0, weight=1)
        self._widget_refs["form"] = form

        ctk.CTkLabel(form, text="Dados do equipamento",
                     font=("Courier New", 15, "bold"), text_color=t["fg_title"],
                     anchor="w").grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 2))

        ctk.CTkLabel(form, text="Cadastre ou atualize as informações.",
                     font=("Courier New", 11), text_color=t["fg_muted"],
                     anchor="w").grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 6))

        corpo = ctk.CTkScrollableFrame(form, fg_color=t["bg_card"], corner_radius=0)
        corpo.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 4))
        corpo.grid_columnconfigure(0, weight=1)
        form.grid_rowconfigure(2, weight=1)
        self._widget_refs["corpo"] = corpo

        campos_frame = ctk.CTkFrame(corpo, fg_color="transparent")
        campos_frame.grid(row=0, column=0, sticky="ew", padx=2)
        campos_frame.grid_columnconfigure(0, weight=1)
        campos_frame.grid_columnconfigure(1, weight=1)

        campos = [
            ("patrimonio", "Patrimônio *", 0, 0),
            ("marca",      "Marca",        0, 1),
            ("modelo",     "Modelo",       1, 0),
            ("serial",     "Serial",       1, 1),
            ("responsavel","Responsável",  2, 0),
            ("setor",      "Setor",        2, 1),
            ("localizacao","Localização",  3, 0),
        ]

        for chave, label, lin, col in campos:
            bloco = ctk.CTkFrame(campos_frame, fg_color="transparent")
            bloco.grid(row=lin, column=col, sticky="ew", padx=4, pady=4)
            bloco.grid_columnconfigure(0, weight=1)

            ctk.CTkLabel(bloco, text=label, font=("Courier New", 11, "bold"),
                         text_color=t["fg_muted"], anchor="w").grid(row=0, column=0, sticky="ew")

            var = tk.StringVar()
            ent = ctk.CTkEntry(bloco, textvariable=var, height=34, corner_radius=7,
                               border_color=t["entry_border"], font=("Courier New", 12),
                               fg_color=t["bg_main"], text_color=t["fg_body"])
            ent.grid(row=1, column=0, sticky="ew", pady=(2, 0))
            self.campos[chave] = var

        sb = ctk.CTkFrame(campos_frame, fg_color="transparent")
        sb.grid(row=3, column=1, sticky="ew", padx=4, pady=4)
        sb.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(sb, text="Status", font=("Courier New", 11, "bold"),
                     text_color=t["fg_muted"], anchor="w").grid(row=0, column=0, sticky="ew")

        self.status = tk.StringVar(value="Em uso")
        self.status_menu = ctk.CTkOptionMenu(
            sb, variable=self.status, values=STATUS, height=34, corner_radius=7,
            fg_color=t["bg_main"], button_color=t["accent"], button_hover_color=t["accent_hover"],
            text_color=t["fg_body"], dropdown_fg_color=t["bg_card"],
            dropdown_text_color=t["fg_body"], font=("Courier New", 12),
        )
        self.status_menu.grid(row=1, column=0, sticky="ew", pady=(2, 0))

        obs_f = ctk.CTkFrame(corpo, fg_color="transparent")
        obs_f.grid(row=1, column=0, sticky="ew", padx=6, pady=(4, 2))
        obs_f.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(obs_f, text="Observações", font=("Courier New", 11, "bold"),
                     text_color=t["fg_muted"], anchor="w").grid(row=0, column=0, sticky="ew")

        self.obs = ctk.CTkTextbox(obs_f, height=65, corner_radius=7,
                                  border_width=1, border_color=t["entry_border"],
                                  font=("Courier New", 12), fg_color=t["bg_main"],
                                  text_color=t["fg_body"])
        self.obs.grid(row=1, column=0, sticky="ew", pady=(3, 0))
        self._widget_refs["obs"] = self.obs

        hist_f = ctk.CTkFrame(corpo, fg_color="transparent")
        hist_f.grid(row=2, column=0, sticky="ew", padx=6, pady=(6, 4))
        hist_f.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hist_f, text="Histórico de movimentações",
                     font=("Courier New", 11, "bold"), text_color=t["fg_muted"],
                     anchor="w").grid(row=0, column=0, sticky="ew")

        self.historico = ctk.CTkTextbox(hist_f, height=100, corner_radius=7,
                                        border_width=1, border_color=t["entry_border"],
                                        font=("Courier New", 11), fg_color=t["bg_main"],
                                        text_color=t["fg_muted"], state="disabled")
        self.historico.grid(row=1, column=0, sticky="ew", pady=(3, 0))
        self._widget_refs["historico"] = self.historico

        botoes = ctk.CTkFrame(form, fg_color=t["bg_card"])
        botoes.grid(row=3, column=0, sticky="ew", padx=14, pady=(6, 14))
        botoes.grid_columnconfigure((0, 1, 2), weight=1)
        self._widget_refs["botoes"] = botoes

        self.btn_novo = ctk.CTkButton(botoes, text="Novo", height=36, corner_radius=8,
                                      fg_color=t["btn_neutral_bg"], hover_color=t["border"],
                                      text_color=t["btn_neutral_fg"], font=("Courier New", 12),
                                      command=self.novo)
        self.btn_novo.grid(row=0, column=0, sticky="ew", padx=(0, 4))

        self.btn_salvar = ctk.CTkButton(botoes, text="Salvar", height=36, corner_radius=8,
                                        fg_color=t["accent"], hover_color=t["accent_hover"],
                                        text_color="#FFFFFF", font=("Courier New", 12, "bold"),
                                        command=self.salvar)
        self.btn_salvar.grid(row=0, column=1, sticky="ew", padx=4)

        self.btn_excluir = ctk.CTkButton(botoes, text="Excluir", height=36, corner_radius=8,
                                         fg_color=t["danger"], hover_color=t["danger_hover"],
                                         text_color="#FFFFFF", font=("Courier New", 12),
                                         command=self.excluir)
        self.btn_excluir.grid(row=0, column=2, sticky="ew", padx=(4, 0))

    def montar_toast(self):
        t = self.T()
        self.toast = ctk.CTkFrame(self, fg_color=t["toast_bg"], corner_radius=10)
        self.toast_label = ctk.CTkLabel(
            self.toast, text="", font=("Courier New", 12, "bold"),
            text_color=t["toast_fg"]
        )
        self.toast_label.pack(padx=18, pady=10)
        self.toast.place_forget()

    def mostrar_toast(self, msg, cor=None):
        t = self.T()
        cor = cor or t["toast_bg"]
        self.toast.configure(fg_color=cor)
        self.toast_label.configure(text=msg)
        self.toast.place(relx=0.5, rely=0.96, anchor="s")
        self.toast.lift()
        if self._toast_job:
            self.after_cancel(self._toast_job)
        self._toast_job = self.after(2800, self.toast.place_forget)

    # ─── paginação ────────────────────────────────────────────────────────────

    def _reset_pagina(self):
        self.pagina_atual = 0
        self.carregar()

    def _pag_prev(self):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self._renderizar_pagina()

    def _pag_next(self):
        total_pags = max(1, (len(self.dados_cache) + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)
        if self.pagina_atual < total_pags - 1:
            self.pagina_atual += 1
            self._renderizar_pagina()

    def _renderizar_pagina(self):
        t = self.T()
        inicio = self.pagina_atual * ROWS_PER_PAGE
        fim    = inicio + ROWS_PER_PAGE
        pagina = self.dados_cache[inicio:fim]

        self.tabela.delete(*self.tabela.get_children())
        for idx, item in enumerate(pagina):
            tag = "par" if idx % 2 == 0 else "impar"
            self.tabela.insert("", "end", iid=item[0], values=item[1:], tags=(tag,))

        self.tabela.tag_configure("par",   background=t["row_even"])
        self.tabela.tag_configure("impar", background=t["row_odd"])

        total_pags = max(1, (len(self.dados_cache) + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)
        self.label_pag.configure(text=f"Página {self.pagina_atual + 1} / {total_pags}")
        self.btn_prev.configure(state="normal" if self.pagina_atual > 0 else "disabled")
        self.btn_next.configure(state="normal" if self.pagina_atual < total_pags - 1 else "disabled")

        total = len(self.dados_cache)
        self.label_contagem.configure(text=f"{total} registro{'s' if total != 1 else ''}")

    def _ordenar(self, coluna):
        dados = [(self.tabela.set(child, coluna), child)
                 for child in self.tabela.get_children("")]
        dados.sort(key=lambda x: x[0].lower())
        for i, (_, child) in enumerate(dados):
            self.tabela.move(child, "", i)

    # ─── filtro ───────────────────────────────────────────────────────────────

    def _aplicar_filtro(self, opcao, cor):
        t = self.T()
        self.filtro_status.set(opcao)
        for op, (btn, c) in self.btns_filtro.items():
            if op == opcao:
                btn.configure(fg_color=c, text_color="#FFFFFF")
            else:
                btn.configure(fg_color=t["btn_neutral_bg"], text_color=t["btn_neutral_fg"])
        self._reset_pagina()

    # ─── dados ────────────────────────────────────────────────────────────────

    def atualizar_resumo(self):
        t = self.T()
        with conectar() as conn:
            total    = conn.execute("SELECT COUNT(*) FROM computadores").fetchone()[0]
            em_uso   = conn.execute("SELECT COUNT(*) FROM computadores WHERE status=?",("Em uso",)).fetchone()[0]
            estoque  = conn.execute("SELECT COUNT(*) FROM computadores WHERE status=?",("Estoque",)).fetchone()[0]
            manut    = conn.execute("SELECT COUNT(*) FROM computadores WHERE status=?",("Manutenção",)).fetchone()[0]

        self.card_total.configure(text=str(total))
        self.card_uso.configure(text=str(em_uso))
        self.card_estoque.configure(text=str(estoque))
        self.card_manutencao.configure(text=str(manut))

    def limpar_busca(self):
        self.busca.set("")
        self.filtro_status.set("Todos")
        for op, (btn, c) in self.btns_filtro.items():
            t = self.T()
            if op == "Todos":
                btn.configure(fg_color=c, text_color="#FFFFFF")
            else:
                btn.configure(fg_color=t["btn_neutral_bg"], text_color=t["btn_neutral_fg"])
        self._reset_pagina()

    def carregar(self):
        busca = f"%{self.busca.get()}%"
        filtro = self.filtro_status.get()

        with conectar() as conn:
            if filtro == "Todos":
                dados = conn.execute("""
                    SELECT id, patrimonio, marca, modelo, responsavel, setor, localizacao, status
                    FROM computadores
                    WHERE (patrimonio LIKE ? OR marca LIKE ? OR modelo LIKE ?
                           OR responsavel LIKE ? OR setor LIKE ? OR localizacao LIKE ? OR status LIKE ?)
                    ORDER BY patrimonio
                """, [busca]*7).fetchall()
            else:
                dados = conn.execute("""
                    SELECT id, patrimonio, marca, modelo, responsavel, setor, localizacao, status
                    FROM computadores
                    WHERE status = ?
                      AND (patrimonio LIKE ? OR marca LIKE ? OR modelo LIKE ?
                           OR responsavel LIKE ? OR setor LIKE ? OR localizacao LIKE ? OR status LIKE ?)
                    ORDER BY patrimonio
                """, [filtro] + [busca]*7).fetchall()

        self.dados_cache = dados
        self._renderizar_pagina()
        self.atualizar_resumo()

    def selecionar(self, event=None):
        item = self.tabela.selection()
        if not item:
            return
        self.id_atual = int(item[0])

        with conectar() as conn:
            dados = conn.execute("""
                SELECT patrimonio, marca, modelo, serial, responsavel, setor, localizacao, status, observacoes
                FROM computadores WHERE id = ?
            """, (self.id_atual,)).fetchone()

            historico = conn.execute("""
                SELECT data_hora, descricao FROM historico
                WHERE computador_id = ? ORDER BY id DESC
            """, (self.id_atual,)).fetchall()

        if not dados:
            return

        chaves = ["patrimonio","marca","modelo","serial","responsavel","setor","localizacao"]
        for i, chave in enumerate(chaves):
            self.campos[chave].set(dados[i] or "")

        self.status.set(dados[7] or "Em uso")
        self.obs.delete("1.0", "end")
        self.obs.insert("1.0", dados[8] or "")

        texto = "\n".join(f"{d} — {desc}" for d, desc in historico)
        self.historico.configure(state="normal")
        self.historico.delete("1.0", "end")
        self.historico.insert("1.0", texto)
        self.historico.configure(state="disabled")

    def novo(self):
        self.id_atual = None
        for var in self.campos.values():
            var.set("")
        self.status.set("Em uso")
        self.obs.delete("1.0", "end")
        self.historico.configure(state="normal")
        self.historico.delete("1.0", "end")
        self.historico.configure(state="disabled")
        self.tabela.selection_remove(self.tabela.selection())

    def salvar(self):
        dados = {k: v.get().strip() for k, v in self.campos.items()}
        dados["status"] = self.status.get()
        dados["observacoes"] = self.obs.get("1.0", "end").strip()

        if not dados["patrimonio"]:
            self.mostrar_toast("⚠  Informe o patrimônio.", self.T()["warning"])
            return

        agora = datetime.now().strftime("%d/%m/%Y %H:%M")

        try:
            with conectar() as conn:
                if self.id_atual is None:
                    cursor = conn.execute("""
                        INSERT INTO computadores
                          (patrimonio,marca,modelo,serial,responsavel,setor,localizacao,status,observacoes,atualizado_em)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (dados["patrimonio"],dados["marca"],dados["modelo"],dados["serial"],
                          dados["responsavel"],dados["setor"],dados["localizacao"],
                          dados["status"],dados["observacoes"],agora))
                    self.id_atual = cursor.lastrowid
                    conn.execute("""
                        INSERT INTO historico (computador_id, data_hora, descricao) VALUES (?,?,?)
                    """, (self.id_atual, agora,
                          f"Cadastrado: resp. {dados['responsavel'] or '-'} | local {dados['localizacao'] or '-'} | {dados['status']}"))
                else:
                    conn.execute("""
                        UPDATE computadores SET
                          patrimonio=?,marca=?,modelo=?,serial=?,responsavel=?,setor=?,
                          localizacao=?,status=?,observacoes=?,atualizado_em=?
                        WHERE id=?
                    """, (dados["patrimonio"],dados["marca"],dados["modelo"],dados["serial"],
                          dados["responsavel"],dados["setor"],dados["localizacao"],
                          dados["status"],dados["observacoes"],agora,self.id_atual))
                    conn.execute("""
                        INSERT INTO historico (computador_id, data_hora, descricao) VALUES (?,?,?)
                    """, (self.id_atual, agora,
                          f"Atualizado: resp. {dados['responsavel'] or '-'} | local {dados['localizacao'] or '-'} | {dados['status']}"))

            self.carregar()
            self.atualizar_resumo()
            self.mostrar_toast("✔  Registro salvo com sucesso.")

        except sqlite3.IntegrityError:
            self.mostrar_toast("✖  Patrimônio já cadastrado.", self.T()["danger"])

    def exportar_excel(self):
        try:
            busca   = f"%{self.busca.get()}%"
            filtro  = self.filtro_status.get()

            with conectar() as conn:
                if filtro == "Todos":
                    dados = conn.execute("""
                        SELECT patrimonio,marca,modelo,serial,responsavel,setor,localizacao,status,observacoes,atualizado_em
                        FROM computadores
                        WHERE (patrimonio LIKE ? OR marca LIKE ? OR modelo LIKE ?
                               OR responsavel LIKE ? OR setor LIKE ? OR localizacao LIKE ? OR status LIKE ?)
                        ORDER BY patrimonio
                    """, [busca]*7).fetchall()
                else:
                    dados = conn.execute("""
                        SELECT patrimonio,marca,modelo,serial,responsavel,setor,localizacao,status,observacoes,atualizado_em
                        FROM computadores
                        WHERE status = ?
                          AND (patrimonio LIKE ? OR marca LIKE ? OR modelo LIKE ?
                               OR responsavel LIKE ? OR setor LIKE ? OR localizacao LIKE ? OR status LIKE ?)
                        ORDER BY patrimonio
                    """, [filtro] + [busca]*7).fetchall()

            if not dados:
                self.mostrar_toast("⚠  Nenhum dado para exportar.", self.T()["warning"])
                return

            RELATORIOS_DIR.mkdir(exist_ok=True)
            ts     = datetime.now().strftime("%Y-%m-%d_%H-%M")
            caminho = RELATORIOS_DIR / f"relatorio_computadores_{ts}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Computadores"

            cabecalhos = ["Patrimônio","Marca","Modelo","Serial","Responsável",
                          "Setor","Localização","Status","Observações","Atualizado em"]
            ws.append(cabecalhos)
            for linha in dados:
                ws.append(list(linha))

            head_fill  = PatternFill("solid", fgColor="1E3A5F")
            head_font  = Font(bold=True, color="FFFFFF", name="Courier New")
            body_font  = Font(name="Courier New", size=10)
            borda      = Border(
                left  =Side(style="thin", color="CBD5E1"),
                right =Side(style="thin", color="CBD5E1"),
                top   =Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )
            alt_fill = PatternFill("solid", fgColor="F1F5F9")

            for cell in ws[1]:
                cell.font      = head_font
                cell.fill      = head_fill
                cell.border    = borda
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if r_idx % 2 == 0 else None
                for cell in row:
                    cell.font      = body_font
                    cell.border    = borda
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if fill:
                        cell.fill = fill

            for col, w in zip("ABCDEFGHIJ", [15,18,22,20,25,18,25,16,35,18]):
                ws.column_dimensions[col].width = w

            ws.row_dimensions[1].height = 22
            ws.freeze_panes  = "A2"
            ws.auto_filter.ref = ws.dimensions

            wb.save(caminho)
            self.mostrar_toast(f"✔  Excel salvo em relatorios/")

        except Exception as erro:
            self.mostrar_toast(f"✖  Erro ao exportar: {erro}", self.T()["danger"])

    def excluir(self):
        if self.id_atual is None:
            self.mostrar_toast("⚠  Selecione um computador.", self.T()["warning"])
            return
        if not messagebox.askyesno("Confirmar exclusão",
                                   "Deseja excluir este computador?\nEsta ação não pode ser desfeita."):
            return
        with conectar() as conn:
            conn.execute("DELETE FROM computadores WHERE id=?",   (self.id_atual,))
            conn.execute("DELETE FROM historico    WHERE computador_id=?", (self.id_atual,))
        self.novo()
        self.carregar()
        self.atualizar_resumo()
        self.mostrar_toast("✔  Registro excluído.")

    # ─── tema ─────────────────────────────────────────────────────────────────

    def alternar_tema(self):
        self.mode = "dark" if self.mode == "light" else "light"
        ctk.set_appearance_mode(self.mode)

        # destroi e reconstrói toda a interface
        for widget in self.winfo_children():
            widget.destroy()

        self._widget_refs.clear()
        self.campos.clear()
        self.id_atual = None

        self.configurar_estilos()
        self.montar_tela()
        self.carregar()
        self.atualizar_resumo()


if __name__ == "__main__":
    app = App()
    app.mainloop()